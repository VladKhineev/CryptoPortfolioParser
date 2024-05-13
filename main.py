import time # время
from LxmlSoup import LxmlSoup # библиотеека по парсеру
import requests # получение данных от сайта
from datetime import datetime # дата
import pandas as pd # Эксель
from openpyxl import load_workbook # Эксель


def fileUpload(name, postName): # Подгружает отсутсвующий файл
    try:
        file = open(f'../CryptoFile/Postcoin/{postName}')
    except IOError as e:
        editFile(name, postName)

        print(f'{postName} отсутствует')
    else:
        file.close()

def editFile(name, postName): # Редактирует файл в нужный для платформы формат
    fileExcel = pd.read_excel(f'../CryptoFile/{name}', usecols='A:B')
    fileExcel.to_excel(f'../CryptoFile/Postcoin/{postName}')

    wb, ws, size, file = openFile(postName)
    number = 2

    ws['A1'] = 'Криптовалюта'
    ws['С1'] = 'Цена'
    ws['D1'] = 'Показатель'
    ws['E1'] = 'Набор данных'

    for i in range(size, number - 1, -1):
        ws[f'A{i}'] = f'{name[:-5]}'
        ws[f'D{i}'] = 'Ежедневная цена'
        ws[f'E{i}'] = 'Факт'
        number += 1

    wb.save(file)
    wb.close()

def editCurrentFile(number, url): # Подгружает Прогноз
    file = f'E:/Vlad/Project/CryptoFile/Postcoin/CurrentPrices.xlsx'

    wb = load_workbook(file)

    ws = wb['Sheet1']

    html = requests.get(url).text
    soup = LxmlSoup(html)
    links = soup.find_all('div', class_='price left')
    for link in links:
        price = link.text()
    price = float(price.replace(' ', ''))

    ws[f'B{number + 1}'] = f'{price}'

    wb.save(file)
    wb.close()

def openFile(nameFile): # открывает файл
    file = f'E:/Vlad/Project/CryptoFile/Postcoin/{nameFile}'

    xl = pd.ExcelFile(file)

    df1 = xl.parse('Sheet1')

    wb = load_workbook(file)

    ws = wb['Sheet1']

    number = len(df1)

    return wb, ws, number, file
def parser(name, wb, ws, number, file, url): # Записывает актуальные данные раз в день
    html = requests.get(url).text
    soup = LxmlSoup(html)
    links = soup.find_all('div', class_='price left')
    for link in links:
        price = link.text()
    price = float(price.replace(' ', ''))
    if str(ws[f'B{number + 1}'].value) != str(datetime.now().date()):
        ws[f'A{number + 2}'] = f'{name}'
        ws[f'B{number + 2}'] = f'{dataNow}'
        ws[f'C{number + 2}'] = f'{price}'
        ws[f'D{number + 2}'] = 'Ежедневная цена'
        ws[f'E{number + 2}'] = 'Факт'

        number += 1

    else:
        ws[f'C{number + 1}'] = f'{price}'


    print(f'{name}: №: {number} Date: {dataNow} Price: {price}')

    wb.save(file)
    wb.close()


try:
    wbBitcoin, wsBitcoin, numberBitcoin, fileBitcoin = openFile('PostBitcoin.xlsx')
    wbEthereum, wsEthereum, numberEthereum, fileEthereum = openFile('PostEthereum.xlsx')
    wbLitecoin, wsLitecoin, numberLitecoin, fileLitecoin = openFile('PostLitecoin.xlsx')
    wbDash, wsDash, numberDash, fileDash = openFile('PostDash.xlsx')
    wbNeo, wsNeo, numberNeo, fileNeo = openFile('PostNeo.xlsx')
    wbXrp, wsXrp, numberXrp, fileXrp = openFile('PostXrp.xlsx')

    while True:
        dataNow = datetime.now().date()
        dataNow = f"{dataNow:%d.%m.%Y}"

        parser('bitcoin', wbBitcoin, wsBitcoin, numberBitcoin, fileBitcoin, 'https://investfunds.ru/indexes/9021/')
        parser('ethereum', wbEthereum, wsEthereum, numberEthereum, fileEthereum, 'https://investfunds.ru/indexes/18539/')
        parser('litecoin', wbLitecoin, wsLitecoin, numberLitecoin, fileLitecoin, 'https://investfunds.ru/indexes/18545/')
        parser('dash', wbDash, wsDash, numberDash, fileDash, 'https://investfunds.ru/indexes/18551/')
        parser('neo', wbNeo, wsNeo, numberNeo, fileNeo, 'https://investfunds.ru/indexes/18555/')
        parser('xrp', wbXrp, wsXrp, numberXrp, fileXrp, 'https://investfunds.ru/indexes/18541/')

        editCurrentFile(1, 'https://investfunds.ru/indexes/9021/')
        editCurrentFile(2, 'https://investfunds.ru/indexes/18551/')
        editCurrentFile(3, 'https://investfunds.ru/indexes/18539/')
        editCurrentFile(4, 'https://investfunds.ru/indexes/18545/')
        editCurrentFile(5, 'https://investfunds.ru/indexes/18555/')
        editCurrentFile(6, 'https://investfunds.ru/indexes/18541/')

        print('#'* 60)
        time.sleep(86400)
except: # В случае отсутвия нужного редактированого файла
    fileUpload('bitcoin.xlsx', 'PostBitcoin.xlsx')
    fileUpload('ethereum.xlsx', 'PostEthereum.xlsx')
    fileUpload('litecoin.xlsx', 'PostLitecoin.xlsx')
    fileUpload('dash.xlsx', 'PostDash.xlsx')
    fileUpload('neo.xlsx', 'PostNeo.xlsx')
    fileUpload('xrp.xlsx', 'PostXrp.xlsx')


    print('Он установился автоматически. Запусите снова.')
